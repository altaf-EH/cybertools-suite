import sys
import os
import json
import csv
from datetime import datetime

import pandas as pd

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.colors import HexColor, white
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.lib.utils import ImageReader

# --------------------------------------------------------
# Copyright (c) 2026 CyberTools
# Licensed under CC BY-NC-ND 4.0 International
# All rights reserved. Commercial use & modification strictly prohibited.
# --------------------------------------------------------

# ============================================================
# CONFIGURATION
# ============================================================

REPORT_DIR = "reports"
SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}

ODD_HOUR_START_MIN = 23 * 60        # 23:00
ODD_HOUR_END_MIN = 5 * 60           # 05:00

COLUMN_ALIASES = {
    "sender_account": [
        "sender_account", "sender account", "from_account", "from account",
        "debit_account", "debit account", "payer_account", "payer account",
        "remitter_account", "remitter account", "source_account", "source account",
        "account_debited", "account debited", "from_ac", "from ac",
        "sender_ac", "sender ac", "debited_account", "debited account",
        "payer_ac", "payer ac", "originating_account", "originating account",
        "debit_ac_no", "debit ac no", "payer", "remitter", "from", "sender",
        "sender_upi_id", "sender upi id", "payer_vpa", "payer vpa",
        "from_a/c", "from a/c", "sender_a/c", "sender a/c", "debit_a/c", "debit a/c",
        "a/c_debited", "a/c debited", "from_a/c_no", "from a/c no",
        "debit_account_no", "debit account no", "debit_account_number",
        "debit account number", "debited_account_no", "debited account no",
        "sender_account_no", "sender account no", "payer_account_no", "payer account no",
        "debit_ac", "debit ac", "from_account_no", "from account no",
        "source_account_no", "source account no", "remitter_account_no",
        "remitter account no", "origin_account", "origin account",
        "originating_ac", "originating ac", "debit_from", "debit from",
        "sender_ac_no", "sender ac no", "payer_ac_no", "payer ac no",
        "payer_account_number", "payer account number",
        "sender_account_number", "sender account number",
        "remitter_ac", "remitter ac", "remitter_ac_no", "remitter ac no",
        "debit_bank_account", "debit bank account", "source_ac", "source ac",
        "source_ac_no", "source ac no", "from_bank_account", "from bank account",
        "paying_account", "paying account", "paying_ac", "paying ac",
    ],

    "receiver_account": [
        "receiver_account", "receiver account", "to_account", "to account",
        "credit_account", "credit account", "payee_account", "payee account",
        "beneficiary_account", "beneficiary account", "destination_account",
        "destination account", "account_credited", "account credited",
        "to_ac", "to ac", "receiver_ac", "receiver ac", "credited_account",
        "credited account", "payee_ac", "payee ac", "terminating_account",
        "terminating account", "credit_ac_no", "credit ac no",
        "payee", "beneficiary", "to", "receiver",
        "receiver_upi_id", "receiver upi id", "payee_vpa", "payee vpa",
        "to_a/c", "to a/c", "receiver_a/c", "receiver a/c", "credit_a/c", "credit a/c",
        "a/c_credited", "a/c credited", "to_a/c_no", "to a/c no",
        "receiver_account_no", "receiver account no", "payee_account_no",
        "payee account no", "beneficiary_account_no", "beneficiary account no",
        "destination_account_no", "destination account no",
        "credited_account_no", "credited account no", "credit_ac", "credit ac",
        "credit_ac_no", "credit ac no", "to_account_no", "to account no",
        "receiver_ac_no", "receiver ac no", "payee_ac_no", "payee ac no",
        "beneficiary_ac", "beneficiary ac", "beneficiary_ac_no", "beneficiary ac no",
        "destination_ac", "destination ac", "destination_ac_no", "destination ac no",
        "credit_bank_account", "credit bank account", "receiving_account",
        "receiving account", "receiving_ac", "receiving ac",
    ],

    "amount": [
        "amount", "amt", "transaction_amount", "transaction amount",
        "txn_amount", "txn amount", "value", "amount_inr", "amount (inr)",
        "amount_rs", "tran_amount", "tran amount", "txn_value", "txn value",
        "debit_amount", "debit amount", "credit_amount", "credit amount",
        "amount_in_rs", "amount in rs", "transaction_value", "transaction value",
        "txn_value_inr", "txn value inr", "amount_inr", "amount inr",
        "rs_amount", "rs amount", "inr_amount", "inr amount",
    ],

    "date": [
        "date", "txn_date", "txn date", "transaction_date", "transaction date",
        "value_date", "value date", "posting_date", "posting date",
        "event_date", "event date",
    ],

    "time": [
        "time", "txn_time", "txn time", "transaction_time", "transaction time",
        "event_time", "event time",
    ],

    "datetime": [
        "datetime", "date_time", "date time", "txn_datetime", "txn datetime",
        "transaction_datetime", "transaction datetime", "timestamp",
        "event_datetime", "event datetime", "event_timestamp", "event timestamp",
        "transaction_timestamp", "transaction timestamp",
    ],

    "transaction_type": [
        "transaction_type", "transaction type", "txn_type", "txn type",
        "type", "dr_cr", "dr/cr", "debit_credit", "debit/credit", "mode",
        "entry_type", "entry type",
    ],

    "channel": [
        "channel", "payment_mode", "payment mode", "transaction_mode",
        "transaction mode", "txn_mode", "txn mode", "payment_channel",
        "payment channel", "instrument", "product_type", "product type",
        "payment_method", "payment method", "payment_type", "payment type",
    ],

    "bank_name": [
        "bank_name", "bank name", "bank", "beneficiary_bank", "beneficiary bank",
        "payer_bank", "payer bank", "ifsc", "ifsc_code", "ifsc code", "branch",
        "bank_ifsc", "bank ifsc", "bank_code", "bank code",
    ],

    "remarks": [
        "remarks", "narration", "description", "particulars", "note",
        "notes", "comments", "purpose",
    ],

    "reference_id": [
        "reference_id", "reference id", "txn_id", "txn id", "transaction_id",
        "transaction id", "utr", "utr_number", "utr number", "rrn",
        "reference_no", "reference no", "ref_no", "ref no",
        "transaction_ref", "transaction ref", "txn_ref", "txn ref",
        "utr_no", "utr no", "rrn_number", "rrn number",
    ],
}

FALLBACK_KEYWORDS = {
    "sender_account": ["payer", "sender", "remitter", "debit", "originat", "from"],
    "receiver_account": ["payee", "beneficiary", "receiver", "credit", "terminat", "destination"],
    "amount": ["amount", "amt", "value"],
    "reference_id": ["utr", "rrn", "ref"],
    "datetime": ["datetime", "timestamp"],
    "date": ["date"],
    "time": ["time"],
    "channel": ["channel", "mode", "instrument"],
    "transaction_type": ["type", "dr_cr", "drcr"],
    "bank_name": ["bank", "ifsc"],
    "remarks": ["narration", "remark", "particular", "description", "note", "purpose"],
}

def get_txn_path():
    """Get transaction file path from command line or user input."""
    positional_args = [
        arg for arg in sys.argv[1:]
        if not arg.startswith("--")
    ]

    if positional_args:
        return positional_args[0].strip().strip('"').strip("'")

    print()
    print("=" * 70)
    print("        FINANCIAL FRAUD & MULE ACCOUNT LINK ANALYZER")
    print("=" * 70)
    print()

    raw = input("Enter transaction statement file path: ").strip()
    return raw.strip('"').strip("'")


def get_output_dir():
    """Read --output flag from command line. Falls back to default."""
    if "--output" in sys.argv:
        idx = sys.argv.index("--output")
        if idx + 1 < len(sys.argv):
            return sys.argv[idx + 1].strip().strip('"').strip("'")
    return REPORT_DIR


def validate_file(file_path):
    """Check if file exists and has supported extension."""
    if not os.path.exists(file_path):
        print(f"[ERROR] File not found: {file_path}")
        sys.exit(1)

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in SUPPORTED_EXTENSIONS:
        print(f"[ERROR] Unsupported file format: {ext}")
        print("Supported formats: .csv, .xlsx, .xls, .tsv, .txt")
        sys.exit(1)

    return True


def check_args():
    if "--credits" in sys.argv:
        print("\n==================================================")
        print("[+] Tool Name: CyberTools FinTrack")
        print("[+] Developer: CyberTools Team (c) 2026")
        print("[+] License  : Creative Commons BY-NC-ND 4.0")
        print("==================================================")
        sys.exit(0)


def format_duration(seconds):
    seconds = int(seconds)
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    remaining_seconds = seconds % 60

    if hours > 0:
        return f"{hours}h {minutes}m {remaining_seconds}s"
    if minutes > 0:
        return f"{minutes}m {remaining_seconds}s"
    return f"{remaining_seconds}s"


def format_currency(value):
    return f"Rs. {value:,.2f}"


def calculate_txn_velocity(records, duration_seconds):
    hours = duration_seconds / 3600 if duration_seconds > 0 else 0

    if hours <= 0:
        rate = float(records)
    else:
        rate = records / hours

    if rate > 20:
        label = "VERY HIGH"
    elif rate > 8:
        label = "HIGH"
    elif rate > 2:
        label = "MODERATE"
    else:
        label = "LOW"

    return label, rate


def calculate_risk(records, counterparties, odd_hour_ratio, pass_through_ratio,
                    same_day_inout_ratio, round_amount_ratio):
    score = 0
    reasons = []

    if records >= 200:
        score += 20
        reasons.append("Very high transaction volume detected.")
    elif records >= 80:
        score += 15
        reasons.append("High transaction volume detected.")
    elif records >= 30:
        score += 10
        reasons.append("Repeated transaction activity detected.")
    elif records >= 10:
        score += 5
        reasons.append("Multiple transaction records detected.")
    else:
        reasons.append("Low transaction volume.")

    if counterparties >= 25:
        score += 15
        reasons.append("Transacted with a very large number of unique accounts.")
    elif counterparties >= 12:
        score += 10
        reasons.append("Transacted with several unique accounts.")
    elif counterparties >= 5:
        score += 5
        reasons.append("Transacted with a moderate number of unique accounts.")
    else:
        reasons.append("Limited number of unique counterparties.")

    if pass_through_ratio >= 0.85:
        score += 30
        reasons.append("Funds pass through almost entirely (classic mule-account signature).")
    elif pass_through_ratio >= 0.6:
        score += 20
        reasons.append("High pass-through of funds - money received is largely transferred out.")
    elif pass_through_ratio >= 0.35:
        score += 10
        reasons.append("Moderate pass-through of funds observed.")
    else:
        reasons.append("Funds are largely retained rather than passed through.")

    if same_day_inout_ratio >= 0.5:
        score += 15
        reasons.append("Majority of active days show same-day in-and-out fund movement (rapid layering).")
    elif same_day_inout_ratio >= 0.2:
        score += 8
        reasons.append("Some same-day in-and-out fund movement detected.")

    if odd_hour_ratio > 0.5:
        score += 10
        reasons.append("Majority of activity occurs during odd hours (11PM-5AM).")
    elif odd_hour_ratio > 0.25:
        score += 6
        reasons.append("Significant odd-hour activity detected.")
    elif odd_hour_ratio > 0.1:
        score += 3
        reasons.append("Some odd-hour activity detected.")
    else:
        reasons.append("Activity occurs mostly during normal hours.")

    if round_amount_ratio >= 0.4:
        score += 10
        reasons.append("High proportion of round-figure transactions detected (possible structuring).")
    elif round_amount_ratio >= 0.2:
        score += 5
        reasons.append("Some round-figure transactions detected.")

    score = min(score, 100)

    if score >= 60:
        risk = "HIGH"
    elif score >= 30:
        risk = "MEDIUM"
    else:
        risk = "LOW"

    return risk, score, reasons


def normalize_header(value):
    text = (
        str(value)
        .strip()
        .lower()
        .replace("-", "_")
        .replace("/", "_")
        .replace(" ", "_")
    )
    text = "".join(ch for ch in text if ch.isalnum() or ch == "_")
    while "__" in text:
        text = text.replace("__", "_")
    return text.strip("_")

def find_column(df, possible_names):
    """Column dhundho chahe uska naam kuch bhi ho."""
    for col in df.columns:
        col_clean = str(col).strip().lower().replace(" ", "_").replace("-", "_")
        for name in possible_names:
            if name in col_clean:
                return col
    return None

def normalize_columns(df):
    """Normalize transaction columns - SMART + FAILPROOF version."""
    
    # 🔥 PEHLE: Smart column detection karo (case insensitive, partial match)
    sender_col = find_column(df, ['sender', 'from', 'payer', 'remitter', 'debit', 'originating', 'source', 'account_debited'])
    receiver_col = find_column(df, ['receiver', 'to', 'payee', 'beneficiary', 'credit', 'destination', 'account_credited'])
    amount_col = find_column(df, ['amount', 'amt', 'value', 'transaction_amount', 'txn_amount', 'inr'])
    date_col = find_column(df, ['date', 'txn_date', 'transaction_date', 'value_date', 'posting_date'])
    time_col = find_column(df, ['time', 'txn_time', 'transaction_time'])
    datetime_col = find_column(df, ['datetime', 'timestamp', 'txn_datetime', 'transaction_datetime'])
    type_col = find_column(df, ['type', 'txn_type', 'transaction_type', 'dr_cr', 'debit_credit'])
    channel_col = find_column(df, ['channel', 'mode', 'payment_mode', 'instrument', 'product_type'])
    bank_col = find_column(df, ['bank', 'bank_name', 'ifsc', 'branch'])
    remarks_col = find_column(df, ['remarks', 'narration', 'description', 'particulars', 'note', 'purpose'])
    ref_col = find_column(df, ['reference', 'ref', 'utr', 'rrn', 'txn_id', 'transaction_id'])
    
    # 🔥 RENAME MAP BANAYO
    rename_map = {}
    
    if sender_col:
        rename_map[sender_col] = "sender_account"
    if receiver_col:
        rename_map[receiver_col] = "receiver_account"
    if amount_col:
        rename_map[amount_col] = "amount"
    if date_col:
        rename_map[date_col] = "date"
    if time_col:
        rename_map[time_col] = "time"
    if datetime_col:
        rename_map[datetime_col] = "datetime"
    if type_col:
        rename_map[type_col] = "transaction_type"
    if channel_col:
        rename_map[channel_col] = "channel"
    if bank_col:
        rename_map[bank_col] = "bank_name"
    if remarks_col:
        rename_map[remarks_col] = "remarks"
    if ref_col:
        rename_map[ref_col] = "reference_id"
    
    # 🔥 FALLBACK: Agar kuch column nahi mila toh purane COLUMN_ALIASES ka use karo
    if not rename_map or "sender_account" not in rename_map or "receiver_account" not in rename_map:
        print("[!] Smart detection incomplete, falling back to alias mapping...")
        # Existing COLUMN_ALIASES logic yahan aayega
        for standard_name, aliases in COLUMN_ALIASES.items():
            for alias in aliases:
                alias_clean = alias.lower().replace(" ", "_").replace("-", "_")
                for col in df.columns:
                    col_clean = str(col).strip().lower().replace(" ", "_").replace("-", "_")
                    if alias_clean == col_clean or alias_clean in col_clean:
                        if standard_name not in rename_map.values():
                            rename_map[col] = standard_name
                            break
    
    # 🔥 FINAL FALLBACK: Agar sender/receiver nahi mile toh first 2 columns assume karo
    if "sender_account" not in rename_map.values() and len(df.columns) >= 2:
        print("[!] Sender column not found, assuming first column is sender...")
        rename_map[df.columns[0]] = "sender_account"
    
    if "receiver_account" not in rename_map.values() and len(df.columns) >= 2:
        print("[!] Receiver column not found, assuming second column is receiver...")
        rename_map[df.columns[1]] = "receiver_account"
    
    if "amount" not in rename_map.values() and len(df.columns) >= 3:
        # Try to find a numeric column
        for col in df.columns:
            if col not in rename_map:
                if pd.api.types.is_numeric_dtype(df[col]):
                    rename_map[col] = "amount"
                    print(f"[!] Assuming numeric column '{col}' is amount...")
                    break
    
    return df.rename(columns=rename_map)


def build_datetime_column(df):
    """Build datetime column - FULLY FAILPROOF."""
    
    # 🔥 Try multiple datetime formats
    date_cols = ['datetime', 'date', 'txn_date', 'transaction_date', 'value_date', 'posting_date']
    time_cols = ['time', 'txn_time', 'transaction_time']
    
    # Find datetime column
    datetime_col = None
    for col in date_cols:
        if col in df.columns:
            datetime_col = col
            break
    
    if datetime_col:
        # Try multiple formats
        formats = ['%d-%m-%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S', '%d-%m-%Y %H:%M', 
                   '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%d/%m/%Y']
        for fmt in formats:
            try:
                df["datetime"] = pd.to_datetime(df[datetime_col], format=fmt, errors='coerce')
                if df["datetime"].notna().sum() > 0:
                    print(f"[+] Date format detected: {fmt}")
                    break
            except:
                continue
        else:
            # Fallback: auto-detect
            df["datetime"] = pd.to_datetime(df[datetime_col], errors='coerce', dayfirst=True)
            print("[+] Date format: auto-detected")
    else:
        print("[ERROR] Could not find date/time columns.")
        print(f"Available columns: {df.columns.tolist()}")
        # 🔥 FALLBACK: Use current time
        from datetime import datetime
        df["datetime"] = datetime.now()
    
    before = len(df)
    df = df.dropna(subset=["datetime"])
    dropped = before - len(df)
    if dropped:
        print(f"[!] Warning: dropped {dropped} row(s) with unreadable date/time.")
    
    return df


def _read_any_table(path):
    """Read any table - with CSV fallback for corrupt Excel files."""
    ext = os.path.splitext(path)[1].lower()
    
    # 🔥 EXCEL FILES
    if ext in (".xlsx", ".xls", ".xlsm"):
        print("[*] Reading Excel file...")
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            sheet_names = wb.sheetnames
            
            if not sheet_names:
                print("[!] No worksheets found. Trying CSV...")
                df = pd.read_csv(path, encoding='utf-8-sig')
                print("[+] Read as CSV")
                return df
            
            print(f"[+] Found sheet: {sheet_names[0]}")
            df = pd.read_excel(path, engine='openpyxl', sheet_name=sheet_names[0])
            df = df.dropna(how='all')
            df = df.dropna(axis=1, how='all')
            print(f"[+] Excel loaded: {len(df)} rows")
            return df
            
        except Exception as e:
            print(f"[!] Excel error: {e}, trying CSV...")
            try:
                df = pd.read_csv(path, encoding='utf-8-sig')
                print("[+] CSV loaded (fallback)")
                return df
            except:
                print("[ERROR] Cannot read file.")
                sys.exit(1)
    
    # 🔥 CSV FILES
    else:
        print("[*] Reading CSV file...")
        try:
            df = pd.read_csv(path, encoding='utf-8-sig')
            df = df.dropna(how='all')
            df = df.dropna(axis=1, how='all')
            print(f"[+] CSV loaded: {len(df)} rows")
            return df
        except Exception as e:
            print(f"[ERROR] Cannot read CSV: {e}")
            sys.exit(1)


def parse_transactions(path):
    """Load and normalize transactions - FULLY FAILPROOF."""
    
    import warnings
    warnings.filterwarnings("ignore")
    
    # 🔥 STEP 1: Read file
    df = _read_any_table(path)
    
    if df is None or df.empty:
        print("[ERROR] No data found.")
        sys.exit(1)
    
    print(f"[+] Raw data: {len(df)} rows, {len(df.columns)} columns")
    print(f"[+] Columns: {df.columns.tolist()}")
    
    # 🔥 STEP 2: Normalize columns
    df = normalize_columns(df)
    print(f"[+] After normalize: {df.columns.tolist()}")
    
    # 🔥 STEP 3: Check required columns - agar nahi mile toh fallback
    required = ["sender_account", "receiver_account", "amount"]
    missing = [c for c in required if c not in df.columns]
    
    if missing:
        print(f"[!] Missing: {missing}")
        print("[!] Using fallback: first 3 columns as Sender, Receiver, Amount")
        cols = df.columns.tolist()
        if len(cols) >= 3:
            rename_map = {}
            rename_map[cols[0]] = "sender_account"
            rename_map[cols[1]] = "receiver_account"
            rename_map[cols[2]] = "amount"
            df = df.rename(columns=rename_map)
            print(f"[+] Fallback mapping: {rename_map}")
    
    # 🔥 STEP 4: Clean amount
    df["amount"] = df["amount"].astype(str).str.replace(',', '').str.replace('Rs', '').str.replace(' ', '').str.strip()
    df["amount"] = pd.to_numeric(df["amount"], errors="coerce")
    df = df.dropna(subset=["amount"])
    df = df[df["amount"] > 0]
    print(f"[+] Valid transactions: {len(df)}")
    
    if df.empty:
        print("[ERROR] No valid transactions after cleaning.")
        sys.exit(1)
    
    # 🔥 STEP 5: Datetime
    if "datetime" in df.columns:
        df["datetime"] = pd.to_datetime(df["datetime"], errors="coerce")
    elif "date" in df.columns and "time" in df.columns:
        combined = df["date"].astype(str).str.strip() + " " + df["time"].astype(str).str.strip()
        df["datetime"] = pd.to_datetime(combined, errors="coerce")
    elif "date" in df.columns:
        df["datetime"] = pd.to_datetime(df["date"], errors="coerce")
    else:
        from datetime import datetime
        df["datetime"] = datetime.now()
        print("[!] No date found, using current time")
    
    df = df.dropna(subset=["datetime"])
    
    # 🔥 STEP 6: Optional columns
    for optional in ("bank_name", "remarks", "reference_id", "transaction_type", "channel"):
        if optional not in df.columns:
            df[optional] = None
    
    # 🔥 STEP 7: Clean strings
    df["sender_account"] = df["sender_account"].astype(str).str.strip()
    df["receiver_account"] = df["receiver_account"].astype(str).str.strip()
    
    df = df[(df["sender_account"] != "") & (df["receiver_account"] != "")]
    
    if df.empty:
        print("[ERROR] No valid records after cleaning.")
        sys.exit(1)
    
    print(f"[+] Final rows: {len(df)}")
    return df.sort_values("datetime").reset_index(drop=True)


# ============================================================
# ACCOUNT ANALYSIS
# ============================================================

def is_odd_hour(dt):
    minute_of_day = dt.hour * 60 + dt.minute
    return minute_of_day >= ODD_HOUR_START_MIN or minute_of_day <= ODD_HOUR_END_MIN


def is_round_amount(amount):
    return amount >= 1000 and amount % 1000 == 0


def analyze_accounts(df):
    """Build a per-account profile: transaction velocity, counterparties,
    pass-through behaviour and layering signals - the core mule-account
    detection logic."""

    account_entries = {}

    for _, row in df.iterrows():
        account_entries.setdefault(row["sender_account"], []).append((row, "OUT", row["receiver_account"]))
        account_entries.setdefault(row["receiver_account"], []).append((row, "IN", row["sender_account"]))

    ranked = sorted(account_entries.items(), key=lambda item: len(item[1]), reverse=True)

    results = []

    for rank, (account, entries) in enumerate(ranked, start=1):
        rows = [e[0] for e in entries]
        counterparties = {}
        for _, _, other in entries:
            counterparties[other] = counterparties.get(other, 0) + 1

        timestamps = [r["datetime"] for r in rows]
        first_seen = min(timestamps)
        last_seen = max(timestamps)
        duration_seconds = (last_seen - first_seen).total_seconds()
        duration_text = format_duration(duration_seconds) if duration_seconds > 0 else "0s"

        odd_hour_count = sum(1 for ts in timestamps if is_odd_hour(ts))
        odd_hour_ratio = odd_hour_count / len(rows) if rows else 0

        velocity_label, rate = calculate_txn_velocity(len(rows), duration_seconds)

        in_entries = [(r, other) for r, d, other in entries if d == "IN"]
        out_entries = [(r, other) for r, d, other in entries if d == "OUT"]

        total_credit = sum(float(r["amount"]) for r, _ in in_entries)
        total_debit = sum(float(r["amount"]) for r, _ in out_entries)
        net_flow = total_credit - total_debit

        pass_through_ratio = (
            min(total_credit, total_debit) / max(total_credit, total_debit)
            if max(total_credit, total_debit) > 0 else 0
        )

        unique_senders = {other for _, other in in_entries}
        unique_receivers = {other for _, other in out_entries}

        amounts = [float(r["amount"]) for r in rows]
        avg_amount = round(sum(amounts) / len(amounts), 2) if amounts else 0
        max_amount = round(max(amounts), 2) if amounts else 0

        round_amount_count = sum(1 for a in amounts if is_round_amount(a))
        round_amount_ratio = round_amount_count / len(rows) if rows else 0

        # Same-day in/out - checks each calendar day the account was active
        # for whether both a credit and a debit happened that day.
        day_directions = {}
        for r, d, _ in entries:
            day = r["datetime"].date()
            day_directions.setdefault(day, set()).add(d)
        same_day_inout_days = sum(1 for dirs in day_directions.values() if len(dirs) == 2)
        total_active_days = len(day_directions) or 1
        same_day_inout_ratio = same_day_inout_days / total_active_days

        channels = {}
        for r in rows:
            ch = str(r["channel"])
            channels[ch] = channels.get(ch, 0) + 1
        channel_text = ", ".join(f"{k} ({v})" for k, v in channels.items())

        top_counterparties = [
            f"{num} ({count})" for num, count in
            sorted(counterparties.items(), key=lambda x: x[1], reverse=True)[:10]
        ]

        risk, risk_score, reasons = calculate_risk(
            len(rows), len(counterparties), odd_hour_ratio,
            pass_through_ratio, same_day_inout_ratio, round_amount_ratio,
        )

        # True funnel signature: money concentrates down to very few
        # destination accounts, regardless of how many debit *transactions*
        # it takes to move it out. (Counting transactions instead of unique
        # receivers misses funnels that sweep out in many small hops.)
        is_funnel = (
            len(in_entries) >= 10 and len(unique_senders) >= 8 and
            0 < len(unique_receivers) <= max(2, int(len(unique_senders) * 0.25))
        )
        if is_funnel:
            risk_score = min(100, risk_score + 15)
            reasons.append(
                "Funnel-account pattern: credits from numerous distinct senders "
                "concentrate down to very few destination accounts."
            )

        # A pure credit-sink: money has piled up from many distinct senders
        # with no matching outward movement in this data window. This is
        # often the accumulation stage of a mule chain (or the funds simply
        # haven't been swept out yet) - either way it warrants a manual look
        # even though its pass-through ratio is zero.
        is_credit_sink = (
            len(in_entries) >= 5 and len(unique_senders) >= 5 and len(out_entries) == 0
        )
        if is_credit_sink:
            risk_score = min(100, risk_score + 20)
            reasons.append(
                "Received funds from numerous distinct senders with no matching "
                "outward activity in this data window - possible accumulation-stage "
                "mule account; recommend manual verification of downstream movement."
            )

        if is_funnel or is_credit_sink:
            if risk_score >= 60:
                risk = "HIGH"
            elif risk_score >= 30:
                risk = "MEDIUM"

        # Closed-loop pair: this account's activity is dominated by one
        # single counterparty (near-exclusive back-and-forth), at real
        # volume. Two accounts that only ever talk to each other - and to
        # nobody else - is itself a distinct red flag, independent of the
        # "limited unique counterparties" penalty in the base score (which
        # was designed to reward diversity of contacts, not to catch this).
        top_counterparty_share = (
            max(counterparties.values()) / len(rows) if rows and counterparties else 0
        )
        is_closed_loop = (
            len(rows) >= 15 and len(counterparties) <= 2 and top_counterparty_share >= 0.85
        )
        if is_closed_loop:
            risk_score = min(100, risk_score + 15)
            reasons.append(
                "Closed-loop pair: nearly all activity is with a single counterparty - "
                "an isolated back-and-forth relationship with no other transaction "
                "history, atypical of a normal account."
            )
            if risk_score >= 60:
                risk = "HIGH"
            elif risk_score >= 30:
                risk = "MEDIUM"

        results.append({
            "rank": rank,
            "account": account,
            "records": len(rows),
            "risk": risk,
            "risk_score": risk_score,
            "first_seen": first_seen.strftime("%d-%m-%Y %H:%M:%S"),
            "last_seen": last_seen.strftime("%d-%m-%Y %H:%M:%S"),
            "duration": duration_text,
            "velocity": velocity_label,
            "rate": rate,
            "unique_counterparties": len(counterparties),
            "in_count": len(in_entries),
            "out_count": len(out_entries),
            "total_credit": round(total_credit, 2),
            "total_debit": round(total_debit, 2),
            "net_flow": round(net_flow, 2),
            "pass_through_ratio": round(pass_through_ratio, 2),
            "same_day_inout_days": same_day_inout_days,
            "avg_amount": avg_amount,
            "max_amount": max_amount,
            "round_amount_count": round_amount_count,
            "odd_hour_count": odd_hour_count,
            "channel_text": channel_text or "N/A",
            "top_counterparties": top_counterparties,
            "is_funnel": is_funnel,
            "is_credit_sink": is_credit_sink,
            "is_closed_loop": is_closed_loop,
            "reasons": reasons,
        })

    return results

# --------------------------------------------------------
# Copyright (c) 2026 CyberTools
# Licensed under CC BY-NC-ND 4.0 International
# All rights reserved. Commercial use & modification strictly prohibited.
# --------------------------------------------------------


def top_account_pairs(df, top_n=15):
    """Link analysis: which pairs of accounts move the most money between
    each other - the raw material for a fund-flow / network graph."""

    pair_counts = {}
    pair_amounts = {}

    for _, row in df.iterrows():
        pair = tuple(sorted([row["sender_account"], row["receiver_account"]]))
        pair_counts[pair] = pair_counts.get(pair, 0) + 1
        pair_amounts[pair] = pair_amounts.get(pair, 0) + float(row["amount"])

    ranked = sorted(pair_amounts.items(), key=lambda x: x[1], reverse=True)[:top_n]
    return [(pair, pair_counts[pair], round(amount, 2)) for pair, amount in ranked]


def find_network_risk_gaps(pairs, account_results, top_n=10):
    """Cross-check the biggest money-flow pairs against each account's own
    risk score. A large fund transfer where one or both endpoints scored
    LOW individually is exactly the kind of thing a per-account-only report
    would bury - surface it explicitly so it doesn't hide in a long list."""

    risk_by_account = {r["account"]: (r["risk"], r["risk_score"]) for r in account_results}
    flagged = []

    for pair, count, amount in pairs[:top_n]:
        risk_a = risk_by_account.get(pair[0], ("LOW", 0))
        risk_b = risk_by_account.get(pair[1], ("LOW", 0))
        if risk_a[0] == "LOW" or risk_b[0] == "LOW":
            flagged.append({
                "account_a": pair[0], "risk_a": risk_a[0], "score_a": risk_a[1],
                "account_b": pair[1], "risk_b": risk_b[0], "score_b": risk_b[1],
                "transactions": count, "total_amount": amount,
            })

    return flagged


def calculate_overall_risk(total_records, unique_accounts, total_amount, account_results):
    if not account_results:
        return "LOW", ["No suspicious transaction activity detected."]

    highest_score = max(r["risk_score"] for r in account_results)
    reasons = []

    if total_records >= 500:
        reasons.append("Large volume of transaction records detected.")
    elif total_records >= 150:
        reasons.append("Moderate-to-high transaction activity detected.")
    else:
        reasons.append("Overall transaction volume is relatively low.")

    if unique_accounts >= 20:
        reasons.append("Large number of unique accounts present in the statement.")
    elif unique_accounts >= 8:
        reasons.append("Multiple unique accounts present in the statement.")

    reasons.append(f"Total value moved across all transactions: {format_currency(total_amount)}.")

    high_pass_through = sum(1 for r in account_results if r["pass_through_ratio"] >= 0.6)
    if high_pass_through:
        reasons.append(f"{high_pass_through} account(s) show a high pass-through ratio (possible mule accounts).")

    funnel_accounts = sum(1 for r in account_results if r["is_funnel"])
    if funnel_accounts:
        reasons.append(f"{funnel_accounts} account(s) show a funnel/structuring pattern.")

    if highest_score >= 60:
        overall_risk = "HIGH"
    elif highest_score >= 30:
        overall_risk = "MEDIUM"
    else:
        overall_risk = "LOW"

    return overall_risk, reasons


# ============================================================
# TERMINAL REPORT
# ============================================================

def print_terminal_report(report_time, total_records, unique_accounts, total_amount,
                           overall_risk, overall_reasons, account_results, pairs, network_flags):

    print()
    print("=" * 100)
    print("            FINANCIAL FRAUD & MULE ACCOUNT ANALYSIS REPORT")
    print("=" * 100)

    print(f"Generated On          : {report_time}")
    print(f"Total Records         : {total_records}")
    print(f"Unique Accounts       : {unique_accounts}")
    print(f"Total Amount Moved    : {format_currency(total_amount)}")
    print(f"Overall Risk          : {overall_risk}")

    print()
    print("Overall Risk Reasons:")
    for reason in overall_reasons:
        print(f"  * {reason}")

    if network_flags:
        print()
        print("=" * 100)
        print("      NETWORK RISK CROSS-CHECK (large money flows involving a LOW-risk account)")
        print("=" * 100)
        for flag in network_flags:
            print(f"  {flag['account_a']} (risk: {flag['risk_a']}/{flag['score_a']}) <-> "
                  f"{flag['account_b']} (risk: {flag['risk_b']}/{flag['score_b']})   "
                  f"txns={flag['transactions']}   total={format_currency(flag['total_amount'])}")

    print()
    print("=" * 100)
    print("                    SUSPICIOUS ACCOUNTS (Top 15)")
    print("=" * 100)

    print(f"{'Rank':<6}{'Account':<20}{'Records':<10}{'Risk':<10}{'Pass-Thru':<12}{'Net Flow':<16}")
    print("-" * 100)

    for result in account_results[:15]:
        print(f"{result['rank']:<6}{result['account']:<20}{result['records']:<10}"
              f"{result['risk']:<10}{result['pass_through_ratio']:<12}{format_currency(result['net_flow']):<16}")
        print(f"      First Seen : {result['first_seen']}")
        print(f"      Last Seen  : {result['last_seen']}")
        print(f"      Credit/Debit: {format_currency(result['total_credit'])} / {format_currency(result['total_debit'])}")
        print()

    print("=" * 100)
    print("                    TOP ACCOUNT PAIRS (LINK ANALYSIS)")
    print("=" * 100)

    for pair, count, amount in pairs[:10]:
        print(f"  {pair[0]} <-> {pair[1]}   txns={count}   total_amount={format_currency(amount)}")

    print()


# ============================================================
# CSV REPORT (flagged accounts - quick import into other systems)
# ============================================================

def generate_csv_report(path, account_results):
    """A flat, minimal CSV of every flagged (MEDIUM/HIGH risk) account -
    the format most banks / fraud desks want for a quick bulk import into
    their own case-management or blocklist systems."""

    fieldnames = [
        "rank", "account", "risk", "risk_score", "records",
        "unique_counterparties", "total_credit", "total_debit", "net_flow",
        "pass_through_ratio", "same_day_inout_days", "is_funnel",
        "first_seen", "last_seen", "top_reason",
    ]

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in account_results:
            if r["risk"] == "LOW":
                continue
            writer.writerow({
                "rank": r["rank"],
                "account": r["account"],
                "risk": r["risk"],
                "risk_score": r["risk_score"],
                "records": r["records"],
                "unique_counterparties": r["unique_counterparties"],
                "total_credit": r["total_credit"],
                "total_debit": r["total_debit"],
                "net_flow": r["net_flow"],
                "pass_through_ratio": r["pass_through_ratio"],
                "same_day_inout_days": r["same_day_inout_days"],
                "is_funnel": r["is_funnel"],
                "first_seen": r["first_seen"],
                "last_seen": r["last_seen"],
                "top_reason": r["reasons"][0] if r["reasons"] else "",
            })


# ============================================================
# JSON REPORT (machine-readable - for handoff to other systems)
# ============================================================

def generate_json_report(path, report_time, total_records, unique_accounts, total_amount,
                          overall_risk, overall_reasons, account_results, pairs, network_flags):

    payload = {
        "tool": "CyberTools FinTrack",
        "report_type": "Financial Fraud & Mule Account Analysis",
        "generated": report_time,
        "summary": {
            "total_records": total_records,
            "unique_accounts": unique_accounts,
            "total_amount_moved": total_amount,
            "overall_risk": overall_risk,
            "overall_risk_reasons": overall_reasons,
        },
        "accounts": account_results,
        "account_pairs": [
            {"account_a": p[0][0], "account_b": p[0][1], "transactions": p[1], "total_amount": p[2]}
            for p in pairs
        ],
        "network_risk_cross_check": network_flags,
    }

    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)


# --------------------------------------------------------
# Copyright (c) 2026 CyberTools
# Licensed under CC BY-NC-ND 4.0 International
# All rights reserved. Commercial use & modification strictly prohibited.
# --------------------------------------------------------

# ============================================================
# XLSX REPORT (for banks / FIU / law-enforcement - sortable, filterable)
# ============================================================

def generate_xlsx_report(path, report_time, total_records, unique_accounts, total_amount,
                          overall_risk, overall_reasons, account_results, pairs, df, network_flags):

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    header_fill = PatternFill(start_color="102A43", end_color="102A43", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    title_font = Font(bold=True, size=13)
    wrap = Alignment(wrap_text=True, vertical="top")

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

    def autosize(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # SHEET 1: SUMMARY
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "FINANCIAL FRAUD & MULE ACCOUNT ANALYSIS REPORT"
    ws["A1"].font = title_font
    ws["A2"] = f"Generated: {report_time}"

    ws["A4"] = "Total Records"
    ws["B4"] = total_records
    ws["A5"] = "Unique Accounts"
    ws["B5"] = unique_accounts
    ws["A6"] = "Total Amount Moved"
    ws["B6"] = total_amount
    ws["A7"] = "Overall Risk"
    ws["B7"] = overall_risk

    for row in ("A4", "A5", "A6", "A7"):
        ws[row].font = Font(bold=True)

    ws["A9"] = "Overall Risk Reasons"
    ws["A9"].font = Font(bold=True)
    for i, reason in enumerate(overall_reasons, start=10):
        ws[f"A{i}"] = f"- {reason}"

    autosize(ws, [65, 18])

    # SHEET 2: SUSPICIOUS ACCOUNTS
    ws2 = wb.create_sheet("Suspicious Accounts")

    columns = [
        "Rank", "Account", "Risk", "Risk Score", "Records", "Unique Counterparties",
        "Credit Txns", "Debit Txns", "Total Credit", "Total Debit", "Net Flow",
        "Pass-Through Ratio", "Same-Day In/Out Days", "Avg Amount", "Max Amount",
        "Round-Figure Txns", "Odd-Hour Records", "Funnel Pattern", "Closed-Loop Pair",
        "First Seen", "Last Seen", "Duration", "Channels Used", "Top Counterparties",
        "Analysis Reasons",
    ]
    ws2.append(columns)
    style_header(ws2)

    for r in account_results:
        ws2.append([
            r["rank"], r["account"], r["risk"], r["risk_score"], r["records"],
            r["unique_counterparties"], r["in_count"], r["out_count"],
            r["total_credit"], r["total_debit"], r["net_flow"], r["pass_through_ratio"],
            r["same_day_inout_days"], r["avg_amount"], r["max_amount"],
            r["round_amount_count"], r["odd_hour_count"], "YES" if r["is_funnel"] else "NO",
            "YES" if r["is_closed_loop"] else "NO",
            r["first_seen"], r["last_seen"], r["duration"], r["channel_text"],
            ", ".join(r["top_counterparties"]), " | ".join(r["reasons"]),
        ])

    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    autosize(ws2, [6, 18, 9, 10, 9, 14, 10, 10, 14, 14, 14, 12, 14, 12, 12, 12, 12, 10, 12, 18, 18, 14, 20, 30, 60])
    ws2.freeze_panes = "A2"
    ws2.page_setup.orientation = "landscape"
    ws2.page_setup.fitToWidth = 1
    ws2.page_setup.fitToHeight = 0
    ws2.sheet_properties.pageSetUpPr.fitToPage = True

    # SHEET 3: ACCOUNT PAIRS (LINK ANALYSIS)
    ws3 = wb.create_sheet("Account Pairs")
    ws3.append(["Account A", "Account B", "Transactions", "Total Amount"])
    style_header(ws3)

    for pair, count, amount in pairs:
        ws3.append([pair[0], pair[1], count, amount])

    autosize(ws3, [20, 20, 14, 18])
    ws3.freeze_panes = "A2"

    # SHEET 3B: NETWORK RISK CROSS-CHECK - large money flows where one or
    # both endpoints scored LOW individually. Meant to catch exactly what a
    # per-account-only view misses: a big transfer hiding behind a quiet
    # individual risk score.
    if network_flags:
        ws3b = wb.create_sheet("Network Risk Cross-Check")
        ws3b.append(["Account A", "Risk A", "Score A", "Account B", "Risk B", "Score B",
                     "Transactions", "Total Amount"])
        style_header(ws3b)
        for flag in network_flags:
            ws3b.append([
                flag["account_a"], flag["risk_a"], flag["score_a"],
                flag["account_b"], flag["risk_b"], flag["score_b"],
                flag["transactions"], flag["total_amount"],
            ])
        autosize(ws3b, [20, 10, 10, 20, 10, 10, 14, 18])
        ws3b.freeze_panes = "A2"

    # SHEET 4: RAW TRANSACTION RECORDS (every original row - for evidence)
    ws4 = wb.create_sheet("Raw Transactions")

    raw_columns = ["Sender Account", "Receiver Account", "Date/Time", "Amount",
                   "Transaction Type", "Channel", "Bank Name", "Reference ID", "Remarks"]
    ws4.append(raw_columns)
    style_header(ws4)

    for _, row in df.iterrows():
        ws4.append([
            row["sender_account"],
            row["receiver_account"],
            row["datetime"].strftime("%d-%m-%Y %H:%M:%S") if pd.notna(row["datetime"]) else "",
            float(row["amount"]) if pd.notna(row["amount"]) else 0,
            row["transaction_type"],
            row["channel"],
            row["bank_name"] if pd.notna(row["bank_name"]) else "",
            row["reference_id"] if pd.notna(row["reference_id"]) else "",
            row["remarks"] if pd.notna(row["remarks"]) else "",
        ])

    autosize(ws4, [20, 20, 20, 14, 16, 14, 18, 20, 30])
    ws4.freeze_panes = "A2"

    wb.save(path)


# ============================================================
# PDF HELPERS (CyberTools standard theme)
# ============================================================

PDF_NAVY = HexColor("#102A43")
PDF_BLUE = HexColor("#168AAD")
PDF_LIGHT_BLUE = HexColor("#EAF4FB")
PDF_BORDER = HexColor("#C9D6E2")
PDF_TEXT = HexColor("#1F2933")
PDF_MUTED = HexColor("#52606D")
PDF_GREEN = HexColor("#16804B")
PDF_ORANGE = HexColor("#D97706")
PDF_RED = HexColor("#C0392B")
PDF_LIGHT_GREEN = HexColor("#EAF7EF")
PDF_LIGHT_ORANGE = HexColor("#FFF4E5")
PDF_LIGHT_RED = HexColor("#FDECEC")

LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "cybertools_logo1.png")


def draw_cybertools_logo(pdf, x, y, width=145):
    if not os.path.exists(LOGO_PATH):
        return
    image = ImageReader(LOGO_PATH)
    image_width, image_height = image.getSize()
    height = width * (image_height / image_width)
    pdf.drawImage(image, x, y - height, width=width, height=height,
                  preserveAspectRatio=True, mask="auto")

# --------------------------------------------------------
# Copyright (c) 2026 CyberTools
# Licensed under CC BY-NC-ND 4.0 International
# All rights reserved. Commercial use & modification strictly prohibited.
# --------------------------------------------------------

def draw_pdf_header(pdf, width, height, page_no):
    draw_cybertools_logo(pdf, 45, height - 18, width=110)

    pdf.setFont("Helvetica-Bold", 7)
    pdf.setFillColor(PDF_MUTED)
    pdf.drawRightString(width - 45, height - 28, f"PAGE {page_no}")

    pdf.setStrokeColor(PDF_BLUE)
    pdf.setLineWidth(1.0)
    pdf.line(45, height - 70, width - 45, height - 70)


def draw_pdf_footer(pdf, width):
    pdf.setStrokeColor(PDF_BORDER)
    pdf.setLineWidth(0.7)
    pdf.line(45, 38, width - 45, 38)

    pdf.setFont("Helvetica", 7)
    pdf.setFillColor(PDF_MUTED)
    pdf.drawString(45, 25, "CyberTools FinTrack")
    pdf.drawRightString(width - 45, 25, "Defensive Fraud Analysis")


def pdf_start_new_page(pdf, width, height, page_no):
    draw_pdf_footer(pdf, width)
    pdf.showPage()
    page_no += 1
    draw_pdf_header(pdf, width, height, page_no)
    return page_no, height - 88


def pdf_ensure_space(pdf, y, required_height, width, height, page_no):
    if y - required_height < 55:
        page_no, y = pdf_start_new_page(pdf, width, height, page_no)
    return page_no, y


def pdf_draw_wrapped_text(pdf, text, x, y, max_width, font="Helvetica", size=8, leading=11):
    pdf.setFont(font, size)
    pdf.setFillColor(PDF_TEXT)

    words = str(text).split()
    lines = []
    current = ""

    for word in words:
        test = word if not current else current + " " + word
        if stringWidth(test, font, size) <= max_width:
            current = test
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)

    for line in lines:
        pdf.drawString(x, y, line)
        y -= leading

    return y


def pdf_draw_section_title(pdf, title, x, y, width):
    pdf.setFillColor(PDF_NAVY)
    pdf.roundRect(x, y - 16, width, 18, 3, fill=1, stroke=0)

    pdf.setFont("Helvetica-Bold", 9)
    pdf.setFillColor(white)
    pdf.drawString(x + 8, y - 11, title)

    return y - 27


def pdf_draw_info_row(pdf, fields, x, y, width, columns=2):
    gap = 7
    cell_width = (width - gap * (columns - 1)) / columns
    cell_height = 34

    rows = [fields[i:i + columns] for i in range(0, len(fields), columns)]

    for row in rows:
        for col_index, (label, value) in enumerate(row):
            cell_x = x + col_index * (cell_width + gap)

            pdf.setFillColor(HexColor("#F7FAFC"))
            pdf.setStrokeColor(PDF_BORDER)
            pdf.setLineWidth(0.6)
            pdf.roundRect(cell_x, y - cell_height, cell_width, cell_height, 3, fill=1, stroke=1)

            pdf.setFont("Helvetica-Bold", 6.8)
            pdf.setFillColor(PDF_MUTED)
            pdf.drawString(cell_x + 7, y - 11, label.upper())

            pdf.setFont("Helvetica-Bold", 8)
            pdf.setFillColor(PDF_TEXT)

            value_text = str(value)
            if len(value_text) > 30:
                value_text = value_text[:27] + "..."

            pdf.drawString(cell_x + 7, y - 25, value_text)

        y -= (cell_height + gap)

    return y


def pdf_draw_risk_badge(pdf, risk, x, y):
    if risk == "HIGH":
        fill, text_color = PDF_LIGHT_RED, PDF_RED
    elif risk == "MEDIUM":
        fill, text_color = PDF_LIGHT_ORANGE, PDF_ORANGE
    else:
        fill, text_color = PDF_LIGHT_GREEN, PDF_GREEN

    badge_width, badge_height = 58, 18

    pdf.setFillColor(fill)
    pdf.setStrokeColor(text_color)
    pdf.roundRect(x, y - badge_height, badge_width, badge_height, 5, fill=1, stroke=1)

    pdf.setFont("Helvetica-Bold", 7.5)
    pdf.setFillColor(text_color)
    pdf.drawCentredString(x + badge_width / 2, y - 12, risk)


# ============================================================
# PDF REPORT
# ============================================================

def generate_pdf_report(path, report_time, total_records, unique_accounts, total_amount,
                         overall_risk, overall_reasons, account_results, pairs, network_flags):

    pdf = canvas.Canvas(path, pagesize=A4)
    width, height = A4
    page_no = 1

    draw_pdf_header(pdf, width, height, page_no)
    y = height - 94

    pdf.setFont("Helvetica-Bold", 14)
    pdf.setFillColor(PDF_TEXT)
    pdf.drawString(45, y, "FINANCIAL FRAUD & MULE ACCOUNT ANALYSIS REPORT")
    y -= 17

    pdf.setFont("Helvetica", 8)
    pdf.setFillColor(PDF_MUTED)
    pdf.drawString(45, y, f"Generated: {report_time}")
    y -= 25

    # Summary cards
    card_gap = 8
    card_width = (width - 90 - card_gap * 3) / 4
    card_height = 54

    summary = [
        ("TOTAL RECORDS", str(total_records)),
        ("UNIQUE ACCOUNTS", str(unique_accounts)),
        ("TOTAL AMOUNT", format_currency(total_amount)),
        ("OVERALL RISK", overall_risk),
    ]

    for index, (label, value) in enumerate(summary):
        card_x = 45 + index * (card_width + card_gap)

        pdf.setFillColor(PDF_LIGHT_BLUE)
        pdf.setStrokeColor(PDF_BORDER)
        pdf.roundRect(card_x, y - card_height, card_width, card_height, 5, fill=1, stroke=1)

        pdf.setFont("Helvetica-Bold", 6.5)
        pdf.setFillColor(PDF_MUTED)
        pdf.drawString(card_x + 8, y - 15, label)

        value_size = 12 if label == "TOTAL AMOUNT" else 15
        pdf.setFont("Helvetica-Bold", value_size)
        if label == "OVERALL RISK":
            pdf.setFillColor(PDF_RED if overall_risk == "HIGH" else
                              PDF_ORANGE if overall_risk == "MEDIUM" else PDF_GREEN)
        else:
            pdf.setFillColor(PDF_NAVY)

        pdf.drawString(card_x + 8, y - 38, value)

    y -= card_height + 18

    y = pdf_draw_section_title(pdf, "OVERALL RISK ASSESSMENT", 45, y, width - 90)

    for reason in overall_reasons:
        page_no, y = pdf_ensure_space(pdf, y, 18, width, height, page_no)
        pdf.setFont("Helvetica", 8)
        pdf.setFillColor(PDF_TEXT)
        pdf.drawString(55, y, "-")
        y = pdf_draw_wrapped_text(pdf, reason, 67, y, width - 122, size=8, leading=10)
        y -= 2

    y -= 8

    if network_flags:
        page_no, y = pdf_ensure_space(pdf, y, 40, width, height, page_no)
        y = pdf_draw_section_title(
            pdf, "NETWORK RISK CROSS-CHECK (large flows involving a LOW-risk account)",
            45, y, width - 90
        )
        for flag in network_flags:
            page_no, y = pdf_ensure_space(pdf, y, 18, width, height, page_no)
            line = (
                f"{flag['account_a']} (risk: {flag['risk_a']}/{flag['score_a']}) <-> "
                f"{flag['account_b']} (risk: {flag['risk_b']}/{flag['score_b']})  "
                f"txns={flag['transactions']}  total={format_currency(flag['total_amount'])}"
            )
            pdf.setFont("Helvetica", 8)
            pdf.setFillColor(PDF_TEXT)
            pdf.drawString(55, y, "-")
            y = pdf_draw_wrapped_text(pdf, line, 67, y, width - 122, size=8, leading=10)
            y -= 2
        y -= 8

    # Suspicious account analysis - starts on a fresh page
    draw_pdf_footer(pdf, width)
    pdf.showPage()
    page_no += 1
    draw_pdf_header(pdf, width, height, page_no)
    y = height - 88

    y = pdf_draw_section_title(pdf, "SUSPICIOUS ACCOUNT ANALYSIS", 45, y, width - 90)

    for result in account_results:
        page_no, y = pdf_ensure_space(pdf, y, 380, width, height, page_no)

        pdf.setFillColor(PDF_NAVY)
        pdf.roundRect(45, y - 28, width - 90, 28, 5, fill=1, stroke=0)

        pdf.setFont("Helvetica-Bold", 11)
        pdf.setFillColor(white)
        pdf.drawString(55, y - 18, f"#{result['rank']}  {result['account']}")

        pdf.setFont("Helvetica-Bold", 7)
        pdf.setFillColor(PDF_MUTED)
        pdf.drawRightString(width - 135, y - 10, "RISK SCORE")

        pdf.setFont("Helvetica-Bold", 11)
        pdf.setFillColor(white)
        pdf.drawRightString(width - 135, y - 22, f"{result['risk_score']}/100")

        pdf_draw_risk_badge(pdf, result["risk"], width - 115, y - 5)

        y -= 40

        fields = [
            ("Records", result["records"]),
            ("Unique Counterparties", result["unique_counterparties"]),
            ("First Seen", result["first_seen"]),
            ("Last Seen", result["last_seen"]),
            ("Duration", result["duration"]),
            ("Transaction Speed", result["velocity"]),
            ("Credit / Debit Txns", f"{result['in_count']} / {result['out_count']}"),
            ("Odd-Hour Records", result["odd_hour_count"]),
            ("Total Credit", format_currency(result["total_credit"])),
            ("Total Debit", format_currency(result["total_debit"])),
            ("Net Flow", format_currency(result["net_flow"])),
            ("Pass-Through Ratio", result["pass_through_ratio"]),
            ("Same-Day In/Out Days", result["same_day_inout_days"]),
            ("Round-Figure Txns", result["round_amount_count"]),
            ("Avg / Max Amount", f"{format_currency(result['avg_amount'])} / {format_currency(result['max_amount'])}"),
            ("Channels Used", result["channel_text"]),
        ]

        page_no, y = pdf_ensure_space(pdf, y, 180, width, height, page_no)
        y = pdf_draw_info_row(pdf, fields, 45, y, width - 90, columns=2)

        page_no, y = pdf_ensure_space(pdf, y, 55, width, height, page_no)
        y = pdf_draw_section_title(pdf, "TOP COUNTERPARTIES", 45, y, width - 90)

        contacts_text = ", ".join(result["top_counterparties"]) if result["top_counterparties"] else "None detected"
        y = pdf_draw_wrapped_text(pdf, contacts_text, 55, y, width - 110, size=8, leading=11)
        y -= 8

        page_no, y = pdf_ensure_space(pdf, y, 55, width, height, page_no)
        y = pdf_draw_section_title(pdf, "ANALYSIS REASONS", 45, y, width - 90)

        for reason in result["reasons"]:
            page_no, y = pdf_ensure_space(pdf, y, 18, width, height, page_no)
            pdf.setFont("Helvetica", 8)
            pdf.setFillColor(PDF_GREEN)
            pdf.drawString(55, y, "-")
            y = pdf_draw_wrapped_text(pdf, reason, 68, y, width - 123, size=8, leading=10)
            y -= 2

        page_no, y = pdf_ensure_space(pdf, y, 18, width, height, page_no)
        y -= 6
        pdf.setStrokeColor(PDF_BORDER)
        pdf.setLineWidth(0.7)
        pdf.line(45, y, width - 45, y)
        y -= 15

    # Link analysis section
    page_no, y = pdf_ensure_space(pdf, y, 200, width, height, page_no)
    y = pdf_draw_section_title(pdf, "TOP ACCOUNT PAIRS (LINK ANALYSIS)", 45, y, width - 90)

    for pair, count, amount in pairs:
        page_no, y = pdf_ensure_space(pdf, y, 16, width, height, page_no)
        pdf.setFont("Helvetica", 8)
        pdf.setFillColor(PDF_TEXT)
        pdf.drawString(55, y, f"{pair[0]} <-> {pair[1]}   txns={count}   total amount={format_currency(amount)}")
        y -= 13

    draw_pdf_footer(pdf, width)
    pdf.save()

# --------------------------------------------------------
# Copyright (c) 2026 CyberTools
# Licensed under CC BY-NC-ND 4.0 International
# All rights reserved. Commercial use & modification strictly prohibited.
# --------------------------------------------------------

# ============================================================
# MAIN
# ============================================================

def main():
    check_args()
    txn_path = get_txn_path()
    
    # Validate file
    validate_file(txn_path)
    
    # Get output directory
    output_dir = get_output_dir()
    os.makedirs(output_dir, exist_ok=True)

    df = parse_transactions(txn_path)

    if df.empty:
        print("[ERROR] No valid transaction records found in the file.")
        print("Please check that the file contains valid transaction data.")
        sys.exit(1)

    total_records = len(df)
    account_results = analyze_accounts(df)
    unique_accounts = len(account_results)
    total_amount = round(float(df["amount"].sum()), 2)
    pairs = top_account_pairs(df)
    network_flags = find_network_risk_gaps(pairs, account_results)

    overall_risk, overall_reasons = calculate_overall_risk(total_records, unique_accounts, total_amount, account_results)

    current_time = datetime.now()
    report_time = current_time.strftime("%d-%m-%Y %H:%M:%S")
    report_file_time = current_time.strftime("%Y%m%d_%H%M%S")

    # SIRF PDF + XLSX - baaki sab HATANA
    pdf_path = os.path.join(output_dir, f"fintrack_report_{report_file_time}.pdf")
    xlsx_path = os.path.join(output_dir, f"fintrack_report_{report_file_time}.xlsx")

    print_terminal_report(report_time, total_records, unique_accounts, total_amount,
                           overall_risk, overall_reasons, account_results, pairs, network_flags)

    # SIRF PDF + XLSX generate karo
    generate_pdf_report(pdf_path, report_time, total_records, unique_accounts, total_amount,
                         overall_risk, overall_reasons, account_results, pairs, network_flags)

    generate_xlsx_report(xlsx_path, report_time, total_records, unique_accounts, total_amount,
                          overall_risk, overall_reasons, account_results, pairs, df, network_flags)

    print()
    print("=" * 70)
    print("Analysis Completed Successfully")
    print("Powered by CyberTools FinTrack")
    print()
    print("PDF Report Saved       :", pdf_path)
    print("XLSX Report Saved      :", xlsx_path)
    print("=" * 70)
    print()


if __name__ == "__main__":
    main()
