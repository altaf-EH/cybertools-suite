import sys
import os
from datetime import datetime, timedelta

import pandas as pd

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.colors import HexColor, white
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.lib.utils import ImageReader


# ============================================================
# CONFIGURATION
# ============================================================

REPORT_DIR = "reports"
# Supported file extensions
SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xls", ".tsv", ".txt"}
ODD_HOUR_START_MIN = 23 * 60        # 23:00
ODD_HOUR_END_MIN = 5 * 60           # 05:00

# Different telecom operators export CDRs with different header names.
# These aliases are normalized onto one internal schema.
COLUMN_ALIASES = {
    "calling_number": [
        "calling_number", "calling number", "a_party", "a party",
        "caller", "calling_no", "calling no", "msisdn", "a_number",
        "a number", "source_number", "source number", "caller_number",
        "caller number", "originating_number", "originating number",
        "originating_party", "originating party", "calling_party",
        "calling party", "from_number", "from number", "from",
        "caller_number", "caller no", "caller_id", "caller id",
        "subscriber_number", "subscriber number", "subscriber_no",
        "subscriber no", "a_party_number", "a party number",
        "a_party_no", "a party no", "originating_number",
        "originating number", "originating_no", "originating no",
        "calling_msisdn", "calling msisdn", "source_msisdn",
        "source msisdn", "caller_msisdn", "caller msisdn",
        "from_msisdn", "from msisdn", "a_msisdn", "a msisdn",
        "calling_subscriber", "calling subscriber",
    ],

    "called_number": [
        "called_number", "called number", "b_party", "b party",
        "callee", "called_no", "called no", "b_number", "b number",
        "destination_number", "destination number", "called_party",
        "called party", "terminating_number", "terminating number",
        "terminating_party", "terminating party", "called_msisdn",
        "called msisdn", "to_number", "to number", "to",
        "called_no", "called id", "destination_no", "destination no",
        "b_party_number", "b party number", "b_party_no", "b party no",
        "terminating_msisdn", "terminating msisdn", "called_msisdn",
        "called msisdn", "receiver_number", "receiver number",
        "receiver_msisdn", "receiver msisdn", "to_msisdn", "to msisdn",
        "b_msisdn", "b msisdn", "called_subscriber", "called subscriber",
    ],

    "date": [
        "date",
        "call_date",
        "call date",
        "start_date",
        "start date",
        "call_start_date",
        "call start date",
        "event_date",
        "event date",
        "transaction_date",
        "transaction date",
    ],

    "time": [
        "time",
        "call_time",
        "call time",
        "start_time",
        "start time",
        "call_start_time",
        "call start time",
        "event_time",
        "event time",
        "transaction_time",
        "transaction time",
    ],

    "datetime": [
        "datetime",
        "date_time",
        "date time",
        "call_datetime",
        "call datetime",
        "start_datetime",
        "start datetime",
        "call_start_datetime",
        "call start datetime",
        "timestamp",
        "event_datetime",
        "event datetime",
        "event_timestamp",
        "event timestamp",
    ],

    "duration": [
        "duration",
        "call_duration",
        "call duration",
        "duration_sec",
        "duration sec",
        "duration_seconds",
        "duration seconds",
        "duration_in_seconds",
        "duration in seconds",
        "talk_time",
        "talk time",
        "talktime",
        "call_length",
        "call length",
        "session_duration",
        "session duration",
    ],

    "call_type": [
        "call_type",
        "call type",
        "type",
        "cdr_type",
        "cdr type",
        "service_type",
        "service type",
        "event_type",
        "event type",
        "call_direction",
        "call direction",
        "traffic_type",
        "traffic type",
    ],

    "tower_id": [
        "tower_id",
        "tower id",
        "tower",
        "towerid",
        "tower_no",
        "tower no",
        "tower_number",
        "tower number",
        "tower_name",
        "tower name",
        "cell_id",
        "cell id",
        "cellid",
        "cell",
        "cell_no",
        "cell no",
        "cell_number",
        "cell number",
        "cell_name",
        "cell name",
        "site_id",
        "site id",
        "siteid",
        "site",
        "lac_cell",
        "lac cell",
        "lac/cell",
        "lac_cell_id",
        "lac cell id",
        "tower_cell_id",
        "tower cell id",
        "tower_cell",
        "tower cell",
        "tower/cell_id",
        "tower/cell id",
        "tower_cell_name",
        "tower cell name",
    ],

    "imei": [
        "imei",
        "imei_number",
        "imei number",
        "a_party_imei",
        "a party imei",
        "calling_imei",
        "calling imei",
        "caller_imei",
        "caller imei",
        "device_imei",
        "device imei",
        "handset_imei",
        "handset imei",
    ],

    "imsi": [
        "imsi",
        "imsi_number",
        "imsi number",
        "a_party_imsi",
        "a party imsi",
        "calling_imsi",
        "calling imsi",
        "caller_imsi",
        "caller imsi",
        "subscriber_imsi",
        "subscriber imsi",
        "sim_imsi",
        "sim imsi",
    ],
}


# ============================================================
# UTILITY FUNCTIONS
# ============================================================

def get_cdr_path():
    """Get CDR file path from command line, or prompt."""
    positional_args = [
        arg for arg in sys.argv[1:]
        if not arg.startswith("--")
    ]

    if positional_args:
        return positional_args[0].strip().strip('"').strip("'")

    print()
    print("=" * 60)
    print("                    CDR ANALYZER")
    print("=" * 60)
    print()

    raw = input("Enter CDR file path: ").strip()
    return raw.strip('"').strip("'")

def get_output_dir():
    """Read --output flag from command line. Falls back to default."""
    if "--output" in sys.argv:
        idx = sys.argv.index("--output")
        if idx + 1 < len(sys.argv):
            return sys.argv[idx + 1].strip().strip('"').strip("'")
    return REPORT_DIR


def validate_file(file_path):
    """Check if file exists and has supported extension."""
    if not os.path.exists(file_path):
        print(f"[ERROR] File not found: {file_path}")
        sys.exit(1)
    
    ext = os.path.splitext(file_path)[1].lower()
    if ext not in SUPPORTED_EXTENSIONS:
        print(f"[ERROR] Unsupported file format: {ext}")
        print("Supported formats: .csv, .xlsx, .xls, .tsv, .txt")
        sys.exit(1)
    
    return True

def check_args():
    if "--credits" in sys.argv:
        print("\n==================================================")
        print("[+] Tool Name: CyberTools CDR Analyzer")
        print("[+] Developer: CyberTools Team (c) 2026")
        print("[+] License  : Creative Commons BY-NC-ND 4.0")
        print("==================================================")
        sys.exit(0)


def format_duration(seconds):
    """Convert seconds into a readable duration."""

    seconds = int(seconds)
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    remaining_seconds = seconds % 60

    if hours > 0:
        return f"{hours}h {minutes}m {remaining_seconds}s"
    if minutes > 0:
        return f"{minutes}m {remaining_seconds}s"
    return f"{remaining_seconds}s"


def calculate_call_rate(records, duration_seconds):
    """Calculate call/session frequency and return label + rate (records/hour)."""

    hours = duration_seconds / 3600 if duration_seconds > 0 else 0

    if hours <= 0:
        rate = float(records)
    else:
        rate = records / hours

    if rate > 20:
        label = "VERY HIGH"
    elif rate > 8:
        label = "HIGH"
    elif rate > 2:
        label = "MODERATE"
    else:
        label = "LOW"

    return label, rate


def calculate_risk(records, contact_count, odd_hour_ratio):
    """Calculate a simple local behavioral risk score for a number."""

    score = 0
    reasons = []

    # Record / call volume
    if records >= 100:
        score += 40
        reasons.append("Very high call/session volume detected.")
    elif records >= 40:
        score += 30
        reasons.append("High call/session volume detected.")
    elif records >= 15:
        score += 20
        reasons.append("Repeated call/session activity detected.")
    elif records >= 5:
        score += 10
        reasons.append("Multiple call/session records detected.")
    else:
        reasons.append("Low call/session volume.")

    # Unique contacts
    if contact_count >= 15:
        score += 25
        reasons.append("Contacted a large number of unique numbers.")
    elif contact_count >= 8:
        score += 15
        reasons.append("Contacted several unique numbers.")
    else:
        reasons.append("Limited number of unique contacts.")

    # Odd-hour activity ratio
    if odd_hour_ratio > 0.5:
        score += 25
        reasons.append("Majority of activity occurs during odd hours (11PM-5AM).")
    elif odd_hour_ratio > 0.25:
        score += 15
        reasons.append("Significant odd-hour activity detected.")
    elif odd_hour_ratio > 0.1:
        score += 5
        reasons.append("Some odd-hour activity detected.")
    else:
        reasons.append("Activity occurs mostly during normal hours.")

    if score >= 60:
        risk = "HIGH"
    elif score >= 30:
        risk = "MEDIUM"
    else:
        risk = "LOW"

    return risk, score, reasons


# ============================================================
# CDR PARSER
# ============================================================

def normalize_header(value):
    """Normalize CDR column headers for flexible alias matching."""
    return (
        str(value)
        .strip()
        .lower()
        .replace("-", "_")
        .replace("/", "_")
        .replace(" ", "_")
    )


def normalize_columns(df):
    rename_map = {}

    normalized_columns = {
        normalize_header(column): column
        for column in df.columns
    }

    for standard_name, aliases in COLUMN_ALIASES.items():

        normalized_aliases = {
            normalize_header(alias)
            for alias in aliases
        }

        for alias in normalized_aliases:
            if alias in normalized_columns:
                original_column = normalized_columns[alias]
                rename_map[original_column] = standard_name
                break

    return df.rename(columns=rename_map)


def build_datetime_column(df):
    if "datetime" in df.columns:
        df["datetime"] = pd.to_datetime(df["datetime"], errors="coerce")
    elif "date" in df.columns and "time" in df.columns:
        combined = df["date"].astype(str).str.strip() + " " + df["time"].astype(str).str.strip()
        df["datetime"] = pd.to_datetime(combined, errors="coerce")
    elif "date" in df.columns:
        df["datetime"] = pd.to_datetime(df["date"], errors="coerce")
    else:
        print("[ERROR] Could not find date/time columns in CDR file.")
        print("Expected either a 'datetime' column, or both 'date' and 'time' columns.")
        sys.exit(1)

    before = len(df)
    df = df.dropna(subset=["datetime"])
    dropped = before - len(df)
    if dropped:
        print(f"[!] Warning: dropped {dropped} row(s) with unreadable date/time.")

    return df


def parse_cdr(cdr_path):
    """Load and normalize a CDR file (CSV or Excel)."""
    ext = os.path.splitext(cdr_path)[1].lower()

    try:
        if ext in (".xlsx", ".xls"):
            df = pd.read_excel(cdr_path)
        else:
            # CSV/TSV - try multiple encodings
            for encoding in ("utf-8-sig", "utf-8", "latin1", "cp1252"):
                try:
                    df = pd.read_csv(cdr_path, encoding=encoding)
                    break
                except Exception:
                    continue
            else:
                raise ValueError("Unable to read CSV file with any encoding")
    except Exception as error:
        print(f"[ERROR] Unable to read CDR file: {error}")
        sys.exit(1)

    df = normalize_columns(df)

    required = ["calling_number", "called_number"]
    missing = [c for c in required if c not in df.columns]

    if missing:
        print(f"[ERROR] Missing required column(s): {missing}")
        print(f"Detected columns: {list(df.columns)}")
        print("Please ensure the file has 'calling_number' and 'called_number' columns.")
        sys.exit(1)

    df = build_datetime_column(df)

    df["duration"] = pd.to_numeric(df["duration"], errors="coerce").fillna(0) if "duration" in df.columns else 0
    df["call_type"] = df["call_type"].fillna("UNKNOWN") if "call_type" in df.columns else "UNKNOWN"

    for optional in ("tower_id", "imei", "imsi"):
        if optional not in df.columns:
            df[optional] = None

    df["calling_number"] = df["calling_number"].astype(str).str.strip()
    df["called_number"] = df["called_number"].astype(str).str.strip()

    # Remove empty rows
    df = df[(df["calling_number"] != "") & (df["called_number"] != "")]

    if df.empty:
        print("[ERROR] No valid CDR records found after cleaning.")
        sys.exit(1)

    return df.sort_values("datetime").reset_index(drop=True)


# ============================================================
# NUMBER ANALYSIS
# ============================================================

def is_odd_hour(dt):
    minute_of_day = dt.hour * 60 + dt.minute
    return minute_of_day >= ODD_HOUR_START_MIN or minute_of_day <= ODD_HOUR_END_MIN


def analyze_numbers(df):
    """Build a per-number profile, similar in spirit to per-IP analysis."""

    number_records = {}

    for _, row in df.iterrows():
        for number, other in ((row["calling_number"], row["called_number"]),
                               (row["called_number"], row["calling_number"])):
            if number not in number_records:
                number_records[number] = []
            number_records[number].append((row, other))

    ranked = sorted(number_records.items(), key=lambda item: len(item[1]), reverse=True)

    results = []

    for rank, (number, entries) in enumerate(ranked, start=1):
        rows = [e[0] for e in entries]
        contacts = {}
        for _, other in entries:
            contacts[other] = contacts.get(other, 0) + 1

        timestamps = [r["datetime"] for r in rows]
        first_seen = min(timestamps)
        last_seen = max(timestamps)
        duration_seconds = (last_seen - first_seen).total_seconds()
        duration_text = format_duration(duration_seconds) if duration_seconds > 0 else "0s"

        odd_hour_count = sum(1 for ts in timestamps if is_odd_hour(ts))
        odd_hour_ratio = odd_hour_count / len(rows) if rows else 0

        speed_label, rate = calculate_call_rate(len(rows), duration_seconds)

        total_talktime = sum(int(r["duration"]) for r in rows)
        avg_duration = round(total_talktime / len(rows), 1) if rows else 0
        short_calls = sum(1 for r in rows if 0 < r["duration"] < 5)

        imeis = sorted({str(r["imei"]) for r in rows if pd.notna(r["imei"])})
        towers = {}
        for r in rows:
            if pd.notna(r["tower_id"]):
                towers[str(r["tower_id"])] = towers.get(str(r["tower_id"]), 0) + 1
        top_tower = max(towers.items(), key=lambda x: x[1])[0] if towers else "N/A"

        call_types = {}
        for r in rows:
            ctype = str(r["call_type"])
            call_types[ctype] = call_types.get(ctype, 0) + 1
        call_type_text = ", ".join(f"{k} ({v})" for k, v in call_types.items())

        risk, risk_score, reasons = calculate_risk(len(rows), len(contacts), odd_hour_ratio)

        if len(imeis) > 1:
            reasons.append(f"Number used {len(imeis)} different IMEIs (possible handset swap).")

        top_contacts = [f"{num} ({count})" for num, count in
                         sorted(contacts.items(), key=lambda x: x[1], reverse=True)[:10]]

        results.append({
            "rank": rank,
            "number": number,
            "records": len(rows),
            "risk": risk,
            "risk_score": risk_score,
            "first_seen": first_seen.strftime("%d-%m-%Y %H:%M:%S"),
            "last_seen": last_seen.strftime("%d-%m-%Y %H:%M:%S"),
            "duration": duration_text,
            "speed": speed_label,
            "rate": rate,
            "unique_contacts": len(contacts),
            "total_talktime": total_talktime,
            "avg_duration": avg_duration,
            "odd_hour_count": odd_hour_count,
            "short_calls": short_calls,
            "imei_count": len(imeis),
            "imeis": imeis,
            "top_tower": top_tower,
            "call_type_text": call_type_text,
            "top_contacts": top_contacts,
            "reasons": reasons,
        })

    return results


def top_contact_pairs(df, top_n=15):
    """Link analysis: which pairs of numbers talk to each other most."""

    pair_counts = {}
    pair_durations = {}

    for _, row in df.iterrows():
        pair = tuple(sorted([row["calling_number"], row["called_number"]]))
        pair_counts[pair] = pair_counts.get(pair, 0) + 1
        pair_durations[pair] = pair_durations.get(pair, 0) + int(row["duration"])

    ranked = sorted(pair_counts.items(), key=lambda x: x[1], reverse=True)[:top_n]
    return [(pair, count, pair_durations[pair]) for pair, count in ranked]


def calculate_overall_risk(total_records, unique_numbers, number_results):

    if not number_results:
        return "LOW", ["No suspicious call activity detected."]

    highest_score = max(r["risk_score"] for r in number_results)
    reasons = []

    if total_records >= 500:
        reasons.append("Large volume of call/session records detected.")
    elif total_records >= 150:
        reasons.append("Moderate-to-high call/session activity detected.")
    else:
        reasons.append("Overall call/session volume is relatively low.")

    if unique_numbers >= 20:
        reasons.append("Large number of unique numbers present in the CDR.")
    elif unique_numbers >= 8:
        reasons.append("Multiple unique numbers present in the CDR.")

    imei_switch_count = sum(1 for r in number_results if r["imei_count"] > 1)
    if imei_switch_count:
        reasons.append(f"{imei_switch_count} number(s) show possible handset (IMEI) swaps.")

    if highest_score >= 60:
        overall_risk = "HIGH"
    elif highest_score >= 30:
        overall_risk = "MEDIUM"
    else:
        overall_risk = "LOW"

    return overall_risk, reasons


# ============================================================
# TERMINAL REPORT
# ============================================================

def print_terminal_report(report_time, total_records, unique_numbers, overall_risk,
                           overall_reasons, number_results, pairs):

    print()
    print("=" * 100)
    print("                         CDR ANALYSIS REPORT")
    print("=" * 100)

    print(f"Generated On          : {report_time}")
    print(f"Total Records         : {total_records}")
    print(f"Unique Numbers        : {unique_numbers}")
    print(f"Overall Risk          : {overall_risk}")

    print()
    print("Overall Risk Reasons:")
    for reason in overall_reasons:
        print(f"  • {reason}")

    print()
    print("=" * 100)
    print("                         SUSPICIOUS NUMBERS (Top 15)")
    print("=" * 100)

    print(f"{'Rank':<6}{'Number':<16}{'Records':<10}{'Risk':<10}{'Contacts':<10}{'Speed':<12}")
    print("-" * 100)

    for result in number_results[:15]:
        print(f"{result['rank']:<6}{result['number']:<16}{result['records']:<10}"
              f"{result['risk']:<10}{result['unique_contacts']:<10}{result['speed']:<12}")
        print(f"      First Seen : {result['first_seen']}")
        print(f"      Last Seen  : {result['last_seen']}")
        print(f"      Duration   : {result['duration']}")
        print(f"      Rate       : {result['rate']:.2f} records/hour")
        print()

    print("=" * 100)
    print("                         TOP CONTACT PAIRS (LINK ANALYSIS)")
    print("=" * 100)

    for pair, count, dur in pairs[:10]:
        print(f"  {pair[0]} <-> {pair[1]}   calls={count}   total_duration={dur}s")

    print()


# ============================================================
# TXT REPORT
# ============================================================

def generate_txt_report(path, report_time, total_records, unique_numbers, overall_risk,
                         overall_reasons, number_results, pairs):

    with open(path, "w", encoding="utf-8") as report:

        report.write("=" * 90 + "\n")
        report.write("CDR ANALYSIS REPORT\n")
        report.write("=" * 90 + "\n")
        report.write(f"Generated: {report_time}\n")
        report.write(f"Total Records   : {total_records}\n")
        report.write(f"Unique Numbers  : {unique_numbers}\n")
        report.write(f"Overall Risk    : {overall_risk}\n\n")

        report.write("Overall Risk Reasons:\n")
        for reason in overall_reasons:
            report.write(f"  - {reason}\n")
        report.write("\n")

        report.write("-" * 90 + "\n")
        report.write("SUSPICIOUS NUMBER ANALYSIS\n")
        report.write("-" * 90 + "\n\n")

        for result in number_results:
            report.write(f"#{result['rank']}  {result['number']}   "
                         f"Risk: {result['risk']} ({result['risk_score']}/90)\n")
            report.write(f"  Records          : {result['records']}\n")
            report.write(f"  First Seen       : {result['first_seen']}\n")
            report.write(f"  Last Seen        : {result['last_seen']}\n")
            report.write(f"  Duration         : {result['duration']}\n")
            report.write(f"  Call Frequency   : {result['speed']} ({result['rate']:.2f}/hour)\n")
            report.write(f"  Unique Contacts  : {result['unique_contacts']}\n")
            report.write(f"  Total Talk-time  : {result['total_talktime']}s\n")
            report.write(f"  Avg Call Duration: {result['avg_duration']}s\n")
            report.write(f"  Odd-Hour Records : {result['odd_hour_count']}\n")
            report.write(f"  Short Calls (<5s): {result['short_calls']}\n")
            report.write(f"  IMEIs Used       : {result['imei_count']}\n")
            report.write(f"  Top Tower/Cell   : {result['top_tower']}\n")
            report.write(f"  Call Types       : {result['call_type_text']}\n")
            report.write(f"  Top Contacts     : {', '.join(result['top_contacts']) if result['top_contacts'] else 'None'}\n")
            report.write("  Analysis Reasons:\n")
            for reason in result["reasons"]:
                report.write(f"    * {reason}\n")
            report.write("\n" + "-" * 90 + "\n\n")

        report.write("TOP CONTACT PAIRS (LINK ANALYSIS)\n")
        report.write("-" * 90 + "\n")
        for pair, count, dur in pairs:
            report.write(f"  {pair[0]} <-> {pair[1]}   calls={count}   total_duration={dur}s\n")

        report.write("\nGenerated by CyberTools CDR Analyzer\n")


# ============================================================
# XLSX REPORT (for police / authorities - sortable, filterable)
# ============================================================

def generate_xlsx_report(path, report_time, total_records, unique_numbers, overall_risk,
                          overall_reasons, number_results, pairs, df):
    """Export the same analysis as a multi-sheet Excel workbook, including raw records."""

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    header_fill = PatternFill(start_color="102A43", end_color="102A43", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    title_font = Font(bold=True, size=13)
    wrap = Alignment(wrap_text=True, vertical="top")

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

    def autosize(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ------------------------------------------------------------------
    # SHEET 1: SUMMARY
    # ------------------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "CDR ANALYSIS REPORT"
    ws["A1"].font = title_font
    ws["A2"] = f"Generated: {report_time}"

    ws["A4"] = "Total Records"
    ws["B4"] = total_records
    ws["A5"] = "Unique Numbers"
    ws["B5"] = unique_numbers
    ws["A6"] = "Overall Risk"
    ws["B6"] = overall_risk

    for row in ("A4", "A5", "A6"):
        ws[row].font = Font(bold=True)

    ws["A8"] = "Overall Risk Reasons"
    ws["A8"].font = Font(bold=True)
    for i, reason in enumerate(overall_reasons, start=9):
        ws[f"A{i}"] = f"- {reason}"

    autosize(ws, [60, 15])

    # ------------------------------------------------------------------
    # SHEET 2: SUSPICIOUS NUMBERS
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("Suspicious Numbers")

    columns = [
        "Rank", "Number", "Risk", "Risk Score", "Records", "Unique Contacts",
        "First Seen", "Last Seen", "Duration", "Call Frequency", "Records/Hour",
        "Odd-Hour Records", "Total Talk-time (s)", "Avg Call Duration (s)",
        "Short Calls (<5s)", "IMEIs Used", "IMEI List", "Top Tower/Cell",
        "Call Types", "Top Contacts", "Analysis Reasons"
    ]
    ws2.append(columns)
    style_header(ws2)

    for r in number_results:
        ws2.append([
            r["rank"], r["number"], r["risk"], r["risk_score"], r["records"],
            r["unique_contacts"], r["first_seen"], r["last_seen"], r["duration"],
            r["speed"], round(r["rate"], 2), r["odd_hour_count"], r["total_talktime"],
            r["avg_duration"], r["short_calls"], r["imei_count"], ", ".join(r["imeis"]),
            r["top_tower"], r["call_type_text"], ", ".join(r["top_contacts"]),
            " | ".join(r["reasons"])
        ])

    for row in ws2.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    autosize(ws2, [6, 14, 9, 10, 9, 14, 18, 18, 14, 14, 12, 16, 16, 18, 16, 10, 40, 14, 25, 45, 60])
    ws2.freeze_panes = "A2"
    ws2.page_setup.orientation = "landscape"
    ws2.page_setup.fitToWidth = 1
    ws2.page_setup.fitToHeight = 0
    ws2.sheet_properties.pageSetUpPr.fitToPage = True

    # ------------------------------------------------------------------
    # SHEET 3: CONTACT PAIRS (LINK ANALYSIS)
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("Contact Pairs")
    ws3.append(["Number A", "Number B", "Calls", "Total Duration (s)"])
    style_header(ws3)

    for pair, count, dur in pairs:
        ws3.append([pair[0], pair[1], count, dur])

    autosize(ws3, [16, 16, 10, 18])
    ws3.freeze_panes = "A2"

    # ------------------------------------------------------------------
    # SHEET 4: RAW CDR RECORDS (every original row - for evidence/verification)
    # ------------------------------------------------------------------
    ws4 = wb.create_sheet("Raw CDR Records")

    raw_columns = ["Calling Number", "Called Number", "Date/Time", "Duration (s)",
                   "Call Type", "Tower/Cell ID", "IMEI", "IMSI"]
    ws4.append(raw_columns)
    style_header(ws4)

    for _, row in df.iterrows():
        ws4.append([
            row["calling_number"],
            row["called_number"],
            row["datetime"].strftime("%d-%m-%Y %H:%M:%S") if pd.notna(row["datetime"]) else "",
            int(row["duration"]) if pd.notna(row["duration"]) else 0,
            row["call_type"],
            row["tower_id"] if pd.notna(row["tower_id"]) else "",
            row["imei"] if pd.notna(row["imei"]) else "",
            row["imsi"] if pd.notna(row["imsi"]) else "",
        ])

    autosize(ws4, [16, 16, 20, 14, 12, 16, 20, 20])
    ws4.freeze_panes = "A2"

    wb.save(path)


# ============================================================
# PDF HELPERS (CyberTools standard theme)
# ============================================================

PDF_NAVY = HexColor("#102A43")
PDF_BLUE = HexColor("#168AAD")
PDF_LIGHT_BLUE = HexColor("#EAF4FB")
PDF_BORDER = HexColor("#C9D6E2")
PDF_TEXT = HexColor("#1F2933")
PDF_MUTED = HexColor("#52606D")
PDF_GREEN = HexColor("#16804B")
PDF_ORANGE = HexColor("#D97706")
PDF_RED = HexColor("#C0392B")
PDF_LIGHT_GREEN = HexColor("#EAF7EF")
PDF_LIGHT_ORANGE = HexColor("#FFF4E5")
PDF_LIGHT_RED = HexColor("#FDECEC")

LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "cybertools_logo1.png")


def draw_cybertools_logo(pdf, x, y, width=145):
    if not os.path.exists(LOGO_PATH):
        return
    image = ImageReader(LOGO_PATH)
    image_width, image_height = image.getSize()
    height = width * (image_height / image_width)
    pdf.drawImage(image, x, y - height, width=width, height=height,
                  preserveAspectRatio=True, mask="auto")


def draw_pdf_header(pdf, width, height, page_no):
    draw_cybertools_logo(pdf, 45, height - 18, width=110)

    pdf.setFont("Helvetica-Bold", 7)
    pdf.setFillColor(PDF_MUTED)
    pdf.drawRightString(width - 45, height - 28, f"PAGE {page_no}")

    pdf.setStrokeColor(PDF_BLUE)
    pdf.setLineWidth(1.0)
    pdf.line(45, height - 70, width - 45, height - 70)


def draw_pdf_footer(pdf, width):
    pdf.setStrokeColor(PDF_BORDER)
    pdf.setLineWidth(0.7)
    pdf.line(45, 38, width - 45, 38)

    pdf.setFont("Helvetica", 7)
    pdf.setFillColor(PDF_MUTED)
    pdf.drawString(45, 25, "CyberTools CDR Analyzer")
    pdf.drawRightString(width - 45, 25, "Defensive Security Analysis")


def pdf_start_new_page(pdf, width, height, page_no):
    draw_pdf_footer(pdf, width)
    pdf.showPage()
    page_no += 1
    draw_pdf_header(pdf, width, height, page_no)
    return page_no, height - 88


def pdf_ensure_space(pdf, y, required_height, width, height, page_no):
    if y - required_height < 55:
        page_no, y = pdf_start_new_page(pdf, width, height, page_no)
    return page_no, y


def pdf_draw_wrapped_text(pdf, text, x, y, max_width, font="Helvetica", size=8, leading=11):
    pdf.setFont(font, size)
    pdf.setFillColor(PDF_TEXT)

    words = str(text).split()
    lines = []
    current = ""

    for word in words:
        test = word if not current else current + " " + word
        if stringWidth(test, font, size) <= max_width:
            current = test
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)

    for line in lines:
        pdf.drawString(x, y, line)
        y -= leading

    return y


def pdf_draw_section_title(pdf, title, x, y, width):
    pdf.setFillColor(PDF_NAVY)
    pdf.roundRect(x, y - 16, width, 18, 3, fill=1, stroke=0)

    pdf.setFont("Helvetica-Bold", 9)
    pdf.setFillColor(white)
    pdf.drawString(x + 8, y - 11, title)

    return y - 27


def pdf_draw_info_row(pdf, fields, x, y, width, columns=2):
    gap = 7
    cell_width = (width - gap * (columns - 1)) / columns
    cell_height = 34

    rows = [fields[i:i + columns] for i in range(0, len(fields), columns)]

    for row in rows:
        for col_index, (label, value) in enumerate(row):
            cell_x = x + col_index * (cell_width + gap)

            pdf.setFillColor(HexColor("#F7FAFC"))
            pdf.setStrokeColor(PDF_BORDER)
            pdf.setLineWidth(0.6)
            pdf.roundRect(cell_x, y - cell_height, cell_width, cell_height, 3, fill=1, stroke=1)

            pdf.setFont("Helvetica-Bold", 6.8)
            pdf.setFillColor(PDF_MUTED)
            pdf.drawString(cell_x + 7, y - 11, label.upper())

            pdf.setFont("Helvetica-Bold", 8)
            pdf.setFillColor(PDF_TEXT)

            value_text = str(value)
            if len(value_text) > 30:
                value_text = value_text[:27] + "..."

            pdf.drawString(cell_x + 7, y - 25, value_text)

        y -= (cell_height + gap)

    return y


def pdf_draw_risk_badge(pdf, risk, x, y):
    if risk == "HIGH":
        fill, text_color = PDF_LIGHT_RED, PDF_RED
    elif risk == "MEDIUM":
        fill, text_color = PDF_LIGHT_ORANGE, PDF_ORANGE
    else:
        fill, text_color = PDF_LIGHT_GREEN, PDF_GREEN

    badge_width, badge_height = 58, 18

    pdf.setFillColor(fill)
    pdf.setStrokeColor(text_color)
    pdf.roundRect(x, y - badge_height, badge_width, badge_height, 5, fill=1, stroke=1)

    pdf.setFont("Helvetica-Bold", 7.5)
    pdf.setFillColor(text_color)
    pdf.drawCentredString(x + badge_width / 2, y - 12, risk)


# ============================================================
# PDF REPORT
# ============================================================

def generate_pdf_report(path, report_time, total_records, unique_numbers, overall_risk,
                         overall_reasons, number_results, pairs):

    pdf = canvas.Canvas(path, pagesize=A4)
    width, height = A4
    page_no = 1

    draw_pdf_header(pdf, width, height, page_no)
    y = height - 94

    pdf.setFont("Helvetica-Bold", 15)
    pdf.setFillColor(PDF_TEXT)
    pdf.drawString(45, y, "CDR ANALYSIS REPORT")
    y -= 17

    pdf.setFont("Helvetica", 8)
    pdf.setFillColor(PDF_MUTED)
    pdf.drawString(45, y, f"Generated: {report_time}")
    y -= 25

    # Summary cards
    card_gap = 8
    card_width = (width - 90 - card_gap * 2) / 3
    card_height = 54

    summary = [("TOTAL RECORDS", str(total_records)),
               ("UNIQUE NUMBERS", str(unique_numbers)),
               ("OVERALL RISK", overall_risk)]

    for index, (label, value) in enumerate(summary):
        card_x = 45 + index * (card_width + card_gap)

        pdf.setFillColor(PDF_LIGHT_BLUE)
        pdf.setStrokeColor(PDF_BORDER)
        pdf.roundRect(card_x, y - card_height, card_width, card_height, 5, fill=1, stroke=1)

        pdf.setFont("Helvetica-Bold", 6.8)
        pdf.setFillColor(PDF_MUTED)
        pdf.drawString(card_x + 9, y - 15, label)

        pdf.setFont("Helvetica-Bold", 15)
        if label == "OVERALL RISK":
            pdf.setFillColor(PDF_RED if overall_risk == "HIGH" else
                              PDF_ORANGE if overall_risk == "MEDIUM" else PDF_GREEN)
        else:
            pdf.setFillColor(PDF_NAVY)

        pdf.drawString(card_x + 9, y - 38, value)

    y -= card_height + 18

    y = pdf_draw_section_title(pdf, "OVERALL RISK ASSESSMENT", 45, y, width - 90)

    for reason in overall_reasons:
        page_no, y = pdf_ensure_space(pdf, y, 18, width, height, page_no)
        pdf.setFont("Helvetica", 8)
        pdf.setFillColor(PDF_TEXT)
        pdf.drawString(55, y, "•")
        y = pdf_draw_wrapped_text(pdf, reason, 67, y, width - 122, size=8, leading=10)
        y -= 2

    y -= 8

    # Suspicious number analysis - starts on a fresh page
    draw_pdf_footer(pdf, width)
    pdf.showPage()
    page_no += 1
    draw_pdf_header(pdf, width, height, page_no)
    y = height - 88

    y = pdf_draw_section_title(pdf, "SUSPICIOUS NUMBER ANALYSIS", 45, y, width - 90)

    for result in number_results:
        page_no, y = pdf_ensure_space(pdf, y, 380, width, height, page_no)

        pdf.setFillColor(PDF_NAVY)
        pdf.roundRect(45, y - 28, width - 90, 28, 5, fill=1, stroke=0)

        pdf.setFont("Helvetica-Bold", 11)
        pdf.setFillColor(white)
        pdf.drawString(55, y - 18, f"#{result['rank']}  {result['number']}")

        pdf.setFont("Helvetica-Bold", 7)
        pdf.setFillColor(PDF_MUTED)
        pdf.drawRightString(width - 135, y - 10, "RISK SCORE")

        pdf.setFont("Helvetica-Bold", 11)
        pdf.setFillColor(white)
        pdf.drawRightString(width - 135, y - 22, f"{result['risk_score']}/90")

        pdf_draw_risk_badge(pdf, result["risk"], width - 115, y - 5)

        y -= 40

        fields = [
            ("Records", result["records"]),
            ("Unique Contacts", result["unique_contacts"]),
            ("First Seen", result["first_seen"]),
            ("Last Seen", result["last_seen"]),
            ("Duration", result["duration"]),
            ("Call Frequency", result["speed"]),
            ("Records / Hour", f"{result['rate']:.2f}"),
            ("Odd-Hour Records", result["odd_hour_count"]),
            ("Total Talk-time", f"{result['total_talktime']}s"),
            ("Avg Call Duration", f"{result['avg_duration']}s"),
            ("Short Calls (<5s)", result["short_calls"]),
            ("IMEIs Used", result["imei_count"]),
            ("Top Tower/Cell", result["top_tower"]),
            ("Call Types", result["call_type_text"] or "N/A"),
        ]

        page_no, y = pdf_ensure_space(pdf, y, 150, width, height, page_no)
        y = pdf_draw_info_row(pdf, fields, 45, y, width - 90, columns=2)

        page_no, y = pdf_ensure_space(pdf, y, 55, width, height, page_no)
        y = pdf_draw_section_title(pdf, "TOP CONTACTS", 45, y, width - 90)

        contacts_text = ", ".join(result["top_contacts"]) if result["top_contacts"] else "None detected"
        y = pdf_draw_wrapped_text(pdf, contacts_text, 55, y, width - 110, size=8, leading=11)
        y -= 8

        page_no, y = pdf_ensure_space(pdf, y, 55, width, height, page_no)
        y = pdf_draw_section_title(pdf, "ANALYSIS REASONS", 45, y, width - 90)

        for reason in result["reasons"]:
            page_no, y = pdf_ensure_space(pdf, y, 18, width, height, page_no)
            pdf.setFont("Helvetica", 8)
            pdf.setFillColor(PDF_GREEN)
            pdf.drawString(55, y, "✓")
            y = pdf_draw_wrapped_text(pdf, reason, 68, y, width - 123, size=8, leading=10)
            y -= 2

        page_no, y = pdf_ensure_space(pdf, y, 18, width, height, page_no)
        y -= 6
        pdf.setStrokeColor(PDF_BORDER)
        pdf.setLineWidth(0.7)
        pdf.line(45, y, width - 45, y)
        y -= 15

    # Link analysis section
    page_no, y = pdf_ensure_space(pdf, y, 200, width, height, page_no)
    y = pdf_draw_section_title(pdf, "TOP CONTACT PAIRS (LINK ANALYSIS)", 45, y, width - 90)

    for pair, count, dur in pairs:
        page_no, y = pdf_ensure_space(pdf, y, 16, width, height, page_no)
        pdf.setFont("Helvetica", 8)
        pdf.setFillColor(PDF_TEXT)
        pdf.drawString(55, y, f"{pair[0]} <-> {pair[1]}   calls={count}   total duration={dur}s")
        y -= 13

    draw_pdf_footer(pdf, width)
    pdf.save()


# ============================================================
# MAIN
# ============================================================

def main():
    check_args()
    cdr_path = get_cdr_path()
    
    # Validate file
    validate_file(cdr_path)
    
    # Get output directory
    output_dir = get_output_dir()
    os.makedirs(output_dir, exist_ok=True)

    df = parse_cdr(cdr_path)

    if df.empty:
        print("[ERROR] No valid CDR records found after cleaning.")
        print("Please check that the CDR file contains valid call records.")
        sys.exit(1)

    total_records = len(df)
    number_results = analyze_numbers(df)
    unique_numbers = len(number_results)
    pairs = top_contact_pairs(df)

    overall_risk, overall_reasons = calculate_overall_risk(total_records, unique_numbers, number_results)

    current_time = datetime.now()
    report_time = current_time.strftime("%d-%m-%Y %H:%M:%S")
    report_file_time = current_time.strftime("%Y%m%d_%H%M%S")

    txt_path = os.path.join(output_dir, f"cdr_report_{report_file_time}.txt")
    pdf_path = os.path.join(output_dir, f"cdr_report_{report_file_time}.pdf")
    xlsx_path = os.path.join(output_dir, f"cdr_report_{report_file_time}.xlsx")

    print_terminal_report(report_time, total_records, unique_numbers, overall_risk,
                           overall_reasons, number_results, pairs)

    generate_txt_report(txt_path, report_time, total_records, unique_numbers, overall_risk,
                         overall_reasons, number_results, pairs)

    generate_pdf_report(pdf_path, report_time, total_records, unique_numbers, overall_risk,
                         overall_reasons, number_results, pairs)

    generate_xlsx_report(xlsx_path, report_time, total_records, unique_numbers, overall_risk,
                          overall_reasons, number_results, pairs, df)

    print()
    print("=" * 70)
    print("Analysis Completed Successfully")
    print("Powered by CyberTools")
    print()
    print("PDF Report Saved  :", pdf_path)
    print("TXT Report Saved  :", txt_path)
    print("XLSX Report Saved :", xlsx_path)
    print("=" * 70)
    print()


if __name__ == "__main__":
    main()